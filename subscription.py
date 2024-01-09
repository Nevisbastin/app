from openpyxl import Workbook, load_workbook
from flask import Flask, render_template, request
import os

# Initialize the Flask application
app = Flask(__name__, template_folder=os.path.abspath('templates'))

# Define the route for the form
@app.route('/update', methods=['GET', 'POST'])
def update_excel():
    if request.method == 'POST':
        email = request.form['email']
        distribution_list = request.form['distribution_list']

        # Check if file exists
        if os.path.isfile('emails.xlsx'):
            wb = load_workbook('emails.xlsx')
            ws = wb.active
        else:
            # Create new file and add headers
            wb = Workbook()
            ws = wb.active
            ws.append(['Email', 'Distribution List'])

        # Update or add data to worksheet
        for row in ws.iter_rows(min_row=2, min_col=1, max_col=1):
            for cell in row:
                if cell.value == email:
                    message = 'Email ID already exists, Please try to add new email ID.'
                    return render_template('error.html', message=message)                    
# Update the email ID in the current row
                    #cell.offset(column=1).value = distribution_list
                    break
            else:
                continue
            break
        else:
            ws.append([email, distribution_list])

        wb.save('emails.xlsx')

        message = 'Email ID has been add, it will take 6 hours for replication.'

        return render_template('success.html', message=message)

    return render_template('form.html')

if __name__ == '__main__':
    app.run(debug=True)
